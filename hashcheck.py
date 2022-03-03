#!/usr/bin/python3

# Se importan los modulos necesarios y se añaden excepciones para los que no vienen instalados por defecto
# La instalacion del modulo vtapi3 en Debian puede requerir de permisos de superusuario para la actualizacion de los repositorios y de la instalacion de python3-pip

import sys
import subprocess
import json
import csv
import argparse
import shutil
import datetime
from time import sleep
from pathlib import Path
import os
import re

try:
    from vtapi3 import VirusTotalAPIFiles, VirusTotalAPIError
except:
    print('\nEl módulo vtapi3 no está instalado. Vamos a intalarlo.\n')
    try:
        subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'vtapi3'])
        from vtapi3 import VirusTotalAPIFiles, VirusTotalAPIError
    except:
        try:
            subprocess.check_call([sys.executable, 'sudo', 'apt', 'update'])
            sleep(2)
            subprocess.check_call([sys.executable, 'sudo', 'apt', '--fix-broken', 'install'])
            sleep(2)
            subprocess.check_call([sys.executable, 'sudo', 'apt', 'install', 'python3-pip'])
            sleep(2)
            subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'vtapi3'])
            sleep(2)
            from vtapi3 import VirusTotalAPIFiles, VirusTotalAPIError
        except:
            pass

try:
    import openpyxl
except:
    print('\nEl módulo openpyxl no está instalado. Vamos a intalarlo.\n')
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl'])
    import openpyxl

from openpyxl.styles import PatternFill, NamedStyle, Font
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from openpyxl.utils import get_column_letter

try:
    from tqdm import tqdm
except:
    print('\nEl módulo tqdm no está instalado. Vamos a intalarlo.\n')
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'tqdm'])
    from tqdm import tqdm

vt_api_files = VirusTotalAPIFiles('Introducir VT API Key')

csv_columns = ['sha256', 'suggested_threat_label', 'malicious', 'suspicious',
               'microsoft', 'trendmicro', 'paloalto', 'kaspersky', 'fortinet',
               'type_extension', 'type_tag', 'sha1', 'md5', 'ssdeep', 'magic',
               'size', 'tags', 'names']

csv_columns2 = ['No matches found']

def logo():
    logo = """

     __                         __                __                         __               
    |  \                       |  \              |  \                       |  \              
    | ▓▓____   ______   _______| ▓▓____   _______| ▓▓____   ______   _______| ▓▓   __         
    | ▓▓    \ |      \ /       \ ▓▓    \ /       \ ▓▓    \ /      \ /       \ ▓▓  /  \        
    | ▓▓▓▓▓▓▓\ \▓▓▓▓▓▓\  ▓▓▓▓▓▓▓ ▓▓▓▓▓▓▓\  ▓▓▓▓▓▓▓ ▓▓▓▓▓▓▓\  ▓▓▓▓▓▓\  ▓▓▓▓▓▓▓ ▓▓_/  ▓▓        
    | ▓▓  | ▓▓/      ▓▓\▓▓    \| ▓▓  | ▓▓ ▓▓     | ▓▓  | ▓▓ ▓▓    ▓▓ ▓▓     | ▓▓   ▓▓         
    | ▓▓  | ▓▓  ▓▓▓▓▓▓▓_\▓▓▓▓▓▓\ ▓▓  | ▓▓ ▓▓_____| ▓▓  | ▓▓ ▓▓▓▓▓▓▓▓ ▓▓_____| ▓▓▓▓▓▓\         
    | ▓▓  | ▓▓\▓▓    ▓▓       ▓▓ ▓▓  | ▓▓\▓▓     \ ▓▓  | ▓▓\▓▓     \\▓▓     \ ▓▓  \▓▓\        
     \▓▓   \▓▓ \▓▓▓▓▓▓▓\▓▓▓▓▓▓▓ \▓▓   \▓▓ \▓▓▓▓▓▓▓\▓▓   \▓▓ \▓▓▓▓▓▓▓ \▓▓▓▓▓▓▓\▓▓   \▓▓        



    """

    print(logo)

def hashcheck():

    with open("hashreport.csv", "a", newline='') as filecsv:
        writer = csv.DictWriter(filecsv, fieldnames=csv_columns)
        writer.writeheader()

    # Crear archivo que almacena hashes no encontrados
    with open('errores404.txt', 'w') as archivo_errores:
        archivo_errores.write('No matches found\n')

    # Sanitizacion de la entrada, regex para sha256, sha1 y md5
    hash_list = []
    with args.filename as f:
        file_item = f.read()

        #Regex sha256
        regex = r'(^[0-9a-fA-F]{64}$)'
        matches = re.finditer(regex, file_item, re.MULTILINE)
        [hash_list.append(match.group())
         for matchNum, match in enumerate(matches, start=1)]

        #Regex sha1
        regex = r'(^[0-9a-fA-F]{40}$)'
        matches = re.finditer(regex, file_item, re.MULTILINE)
        [hash_list.append(match.group())
         for matchNum, match in enumerate(matches, start=1)]

        #Regex md5
        regex = r'(^[0-9a-fA-F]{32}$)'
        matches = re.finditer(regex, file_item, re.MULTILINE)
        [hash_list.append(match.group())
         for matchNum, match in enumerate(matches, start=1)]

    for hash in hash_list:

        try:
            result = vt_api_files.get_report(hash)
        except VirusTotalAPIError as err:
            print(err, err.err_code)
        else:

            if vt_api_files.get_last_http_error() == vt_api_files.HTTP_OK:

                try:
                    result = json.loads(result)
                    result = json.dumps(result, sort_keys=False, indent=4)
                    result = json.loads(result)
                    result = result['data']
                    result = result['attributes']

                    # Positivos (malicious y suspicious)
                    positivos = result['last_analysis_stats']
                    malicious = positivos['malicious']
                    suspicious = positivos['suspicious']

                    # Resultados de los antivirus que nos interesan
                    engines = result['last_analysis_results']
                    microsoft = engines['Microsoft']
                    microsoft = microsoft['category']
                    trendmicro = engines['TrendMicro']
                    trendmicro = trendmicro['category']
                    paloalto = engines['Paloalto']
                    paloalto = paloalto['category']
                    kaspersky = engines['Kaspersky']
                    kaspersky = kaspersky['category']
                    fortinet = engines['Fortinet']
                    fortinet = fortinet['category']
                    suggested_threat_label = result['popular_threat_classification']
                    suggested_threat_label = suggested_threat_label['suggested_threat_label']

                    # Se insertan los positivos en el diccionario
                    result['malicious'] = (malicious)
                    result['suspicious'] = (suspicious)

                    # Se insertan en el diccionario los resultados de los engines que nos interesan
                    result['microsoft'] = (microsoft)
                    result['trendmicro'] = (trendmicro)
                    result['paloalto'] = (paloalto)
                    result['kaspersky'] = (kaspersky)
                    result['fortinet'] = (fortinet)

                except KeyError:
                    pass

                # Creamos un nuevo diccionario para reorganizar los columans en el orden que queremos
                try:
                    result_rearrange = {}
                    sha256 = result['sha256']
                    result_rearrange['sha256'] = (sha256)
                    result_rearrange['suggested_threat_label'] = (suggested_threat_label)
                    result_rearrange['malicious'] = (malicious)
                    result_rearrange['suspicious'] = (suspicious)
                    result_rearrange['microsoft'] = (microsoft)
                    result_rearrange['trendmicro'] = (trendmicro)
                    result_rearrange['paloalto'] = (paloalto)
                    result_rearrange['kaspersky'] = (kaspersky)
                    result_rearrange['fortinet'] = (fortinet)
                    type_extension = result['type_extension']
                    result_rearrange['type_extension'] = (type_extension)
                    type_tag = result['type_tag']
                    result_rearrange['type_tag'] = (type_tag)
                    sha1 = result['sha1']
                    result_rearrange['sha1'] = (sha1)
                    md5 = result['md5']
                    result_rearrange['md5'] = (md5)
                    ssdeep = result['ssdeep']
                    result_rearrange['ssdeep'] = (ssdeep)
                    magic = result['magic']
                    result_rearrange['magic'] = (magic)
                    size = result['size']
                    result_rearrange['size'] = (size)
                    tags = result['tags']
                    result_rearrange['tags'] = (tags)
                    names = result['names']
                    result_rearrange['names'] = (names)

                except KeyError:
                    pass

                with open("hashreport.csv", "a", newline='') as filecsv:
                    writer = csv.DictWriter(filecsv, fieldnames=csv_columns)
                    writer.writerow(result_rearrange)


                print('\nComprobando ' + hash)

                for i in tqdm(range(100)):
                    sleep(0.015)

            else:

                if str(vt_api_files.get_last_http_error()) == '404':
                    print('\nHTTP Error [' + str(vt_api_files.get_last_http_error()) + ']')
                    print(f'No matches found for {hash}')
                    with open("errores404.txt", "a", newline='') as f:
                        f.writelines(hash)
                        f.writelines('\n')
                elif str(vt_api_files.get_last_http_error()) == '429':
                    print('\nHTTP Error [' + str(vt_api_files.get_last_http_error()) + ']')
                    print('Creditos API agotados')
                    break

                else:
                    print('\nHTTP Error [' + str(vt_api_files.get_last_http_error()) + ']')


def report():
    csv_data = []
    with open('hashreport.csv') as file_obj:
        reader = csv.reader(file_obj)
        for row in reader:
            csv_data.append(row)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for row in csv_data:
        sheet.append(row)

    # Ajustar tamaño de las columnas

    MIN_WIDTH = 10

    for i, column_cells in enumerate(sheet.columns, start=1):
        width = (
            length
            if (length := max(len(str(cell_value) if (cell_value := cell.value) is not None else "")
                              for cell in column_cells)) >= MIN_WIDTH
            else MIN_WIDTH
        )
        sheet.column_dimensions[get_column_letter(i)].width = width

    # Primera fila en negrita

    header = NamedStyle(name="header")
    header.font = Font(bold=True)
    header_row = sheet[1]
    for cell in header_row:
        cell.style = header

    # Pintar en rojo IPs reportadas

    red_background = PatternFill(bgColor="ff4040", fill_type="solid")
    diff_style = DifferentialStyle(fill=red_background)
    rule = Rule(type="expression", dxf=diff_style, stopIfTrue=True)
    rule.formula = ['int($C1)>3']
    sheet.conditional_formatting.add("A1:R1000", rule)

    # Congelar primera fila

    sheet.freeze_panes = "A2"

    # Añadir filtros

    sheet.auto_filter.ref = "A1:R1000"

    # Nombre sheet

    sheet.title = 'VT_hashes_results'
    workbook.save('hashreport.xlsx')
    sleep(1)

def report_errores():
    csv_data = []
    with open('hashesNoEncontrados.csv', 'w', newline='') as file_obj, open('errores404.txt', 'r') as f:
        for i in f.readlines():
            writer = csv.writer(file_obj)
            writer.writerow([i.rstrip()])

    with open('hashesNoEncontrados.csv') as file_obj:
        reader = csv.reader(file_obj)
        for row in reader:
            csv_data.append(row)


        # Excepcion por si el excel no ha sido creado todavia
        try:
            workbook = openpyxl.load_workbook('hashreport.xlsx')
            sheet = workbook.create_sheet()
            for row in csv_data:
                sheet.append(row)
        except:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            for row in csv_data:
                sheet.append(row)

        # Ajustar tamaño de las columnas

        MIN_WIDTH = 10
        for i, column_cells in enumerate(sheet.columns, start=1):
            width = (
                length
                if (length := max(len(str(cell_value) if (cell_value := cell.value) is not None else "")
                                  for cell in column_cells)) >= MIN_WIDTH
                else MIN_WIDTH
            )
            sheet.column_dimensions[get_column_letter(i)].width = width

        # Nombre sheet

        sheet.title = 'No matches found'
        workbook.save('hashreport.xlsx')
        sleep(1)

def clean():
    fecha = (datetime.datetime.now()).strftime('%Y-%m-%d-%H-%M')

    archivos = ['hashreport.csv', 'hashreport.xlsx', 'hashesNoEncontrados.csv']

    ruta_inicial = Path('./')
    ruta_final = Path(f'{ruta_inicial}/report_{fecha}')

    # Se crea directorio que va a contener el report
    os.mkdir(ruta_final)

    # Se renombran y mueven los archivos al directorio
    for i in archivos:
        nombre_final = f'{fecha}_{i}'

        shutil.move(ruta_inicial.joinpath(i).resolve(), ruta_final.joinpath(nombre_final).resolve())

    # Se elimina el txt
    os.remove('errores404.txt')

# Opciones
parser = argparse.ArgumentParser()
parser.add_argument('filename', type=argparse.FileType('r'))
args = parser.parse_args()

if args.filename:
    logo()
    hashcheck()
    report()
    report_errores()
    clean()
